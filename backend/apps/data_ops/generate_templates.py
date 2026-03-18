#!/usr/bin/env python3
"""
Generate Excel import templates for ChangePoint data migration.

Creates one Excel file per data domain with:
- Column headers matching ERP model fields
- Data validation (dropdowns) where applicable
- Example row showing expected format
- Instructions sheet

Usage: python generate_templates.py
Output: ../import_templates/*.xlsx
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "import_templates")
os.makedirs(OUTPUT_DIR, exist_ok=True)

HEADER_FILL = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
EXAMPLE_FILL = PatternFill(start_color="F0F9FF", end_color="F0F9FF", fill_type="solid")
REQUIRED_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def style_header(ws, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col)].width = 20


def style_example(ws, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=2, column=col)
        cell.fill = EXAMPLE_FILL
        cell.border = THIN_BORDER


def add_instructions(wb, instructions):
    ws = wb.create_sheet("Instructions", 0)
    ws.column_dimensions["A"].width = 80
    for i, line in enumerate(instructions, 1):
        cell = ws.cell(row=i, column=1, value=line)
        if i == 1:
            cell.font = Font(bold=True, size=14)
        elif line.startswith("•"):
            cell.font = Font(size=11)
        else:
            cell.font = Font(size=11, italic=True)


def create_employees():
    wb = Workbook()
    add_instructions(wb, [
        "IMPORT EMPLOYÉS",
        "",
        "• Remplir l'onglet 'Employes' avec les données ChangePoint",
        "• Colonnes obligatoires marquées avec * dans le header",
        "• username = identifiant ChangePoint (sera le login)",
        "• email = courriel Microsoft (pour SSO Entra ID)",
        "• role = un parmi: EMPLOYEE, PM, PROJECT_DIRECTOR, BU_DIRECTOR, FINANCE, DEPT_ASSISTANT, PROPOSAL_MANAGER, ADMIN",
        "",
        "Format attendu:",
        "• Dates: AAAA-MM-JJ (ex: 2026-03-18)",
        "• Texte: UTF-8 (accents supportés)",
    ])

    ws = wb.create_sheet("Employes")
    headers = [
        "username *", "email *", "first_name *", "last_name *",
        "role *", "business_unit", "position_profile",
        "language_preference"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, len(headers))

    # Example row
    example = [
        "jean.tremblay", "jean.tremblay@provencher-roy.com",
        "Jean", "Tremblay", "PM", "Architecture",
        "Architecte", "fr"
    ]
    for col, val in enumerate(example, 1):
        ws.cell(row=2, column=col, value=val)
    style_example(ws, len(headers))

    wb.save(os.path.join(OUTPUT_DIR, "01_employes.xlsx"))
    print("✓ 01_employes.xlsx")


def create_clients():
    wb = Workbook()
    add_instructions(wb, [
        "IMPORT CLIENTS & CONTACTS",
        "",
        "• Onglet 'Clients' : un client par ligne",
        "• Onglet 'Contacts' : un contact par ligne, lié au client par 'client_ref'",
        "• Onglet 'Adresses' : une adresse par ligne, liée au client par 'client_ref'",
        "• client_ref = identifiant unique du client (code ChangePoint ou nom)",
        "",
        "Format attendu:",
        "• status: active, inactive, archived",
        "• language_preference: fr, en",
        "• is_billing: oui, non",
    ])

    # Clients sheet
    ws = wb.create_sheet("Clients")
    headers = [
        "client_ref *", "name *", "alias", "legal_entity",
        "sector", "status", "payment_terms_days",
        "associe_en_charge", "notes"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, len(headers))
    example = [
        "CLI-001", "Ville de Montréal", "VDM", "Corporation municipale",
        "Public", "active", 30, "Pierre Lavoie", "Client depuis 2015"
    ]
    for col, val in enumerate(example, 1):
        ws.cell(row=2, column=col, value=val)
    style_example(ws, len(headers))

    # Contacts sheet
    ws2 = wb.create_sheet("Contacts")
    headers2 = [
        "client_ref *", "name *", "role", "email",
        "phone", "language_preference"
    ]
    for col, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=col, value=h)
    style_header(ws2, len(headers2))
    example2 = [
        "CLI-001", "Marie Dupont", "Chef de projet client",
        "marie.dupont@ville.montreal.qc.ca", "514-555-0101", "fr"
    ]
    for col, val in enumerate(example2, 1):
        ws2.cell(row=2, column=col, value=val)
    style_example(ws2, len(headers2))

    # Addresses sheet
    ws3 = wb.create_sheet("Adresses")
    headers3 = [
        "client_ref *", "address_line_1 *", "address_line_2",
        "city *", "province", "postal_code *", "country",
        "is_billing", "is_primary"
    ]
    for col, h in enumerate(headers3, 1):
        ws3.cell(row=1, column=col, value=h)
    style_header(ws3, len(headers3))
    example3 = [
        "CLI-001", "275 rue Notre-Dame Est", "Bureau 300",
        "Montréal", "Québec", "H2Y 1C6", "Canada", "oui", "oui"
    ]
    for col, val in enumerate(example3, 1):
        ws3.cell(row=2, column=col, value=val)
    style_example(ws3, len(headers3))

    wb.save(os.path.join(OUTPUT_DIR, "02_clients.xlsx"))
    print("✓ 02_clients.xlsx")


def create_projects():
    wb = Workbook()
    add_instructions(wb, [
        "IMPORT PROJETS, PHASES & WBS",
        "",
        "• Onglet 'Projets' : un projet par ligne",
        "• Onglet 'Phases' : une phase par ligne, liée au projet par 'project_code'",
        "• Onglet 'WBS' : un élément WBS par ligne, lié à la phase par 'phase_ref'",
        "• client_ref doit correspondre à un client_ref du fichier 02_clients.xlsx",
        "• pm_username doit correspondre à un username du fichier 01_employes.xlsx",
        "",
        "Format attendu:",
        "• contract_type: FORFAITAIRE, CONSORTIUM, CO_DEV, CONCEPTION_CONSTRUCTION",
        "• status: ACTIVE, ON_HOLD, COMPLETED, CANCELLED",
        "• billing_mode: FORFAIT, HORAIRE",
        "• Montants: nombre décimal (ex: 50000.00)",
        "• Heures: nombre décimal (ex: 120.5)",
    ])

    # Projects sheet
    ws = wb.create_sheet("Projets")
    headers = [
        "project_code *", "name *", "client_ref *",
        "contract_type *", "status", "business_unit",
        "start_date", "end_date",
        "pm_username", "associe_en_charge_username",
        "is_internal"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, len(headers))
    example = [
        "PRJ-2024-015", "Musée Maritime Québec", "CLI-001",
        "FORFAITAIRE", "ACTIVE", "Architecture",
        "2024-06-01", "2026-12-31",
        "jean.tremblay", "pierre.lavoie", "non"
    ]
    for col, val in enumerate(example, 1):
        ws.cell(row=2, column=col, value=val)
    style_example(ws, len(headers))

    # Phases sheet
    ws2 = wb.create_sheet("Phases")
    headers2 = [
        "project_code *", "phase_ref *", "name *",
        "client_facing_label", "phase_type",
        "billing_mode", "order",
        "start_date", "end_date",
        "budgeted_hours", "budgeted_cost",
        "is_mandatory"
    ]
    for col, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=col, value=h)
    style_header(ws2, len(headers2))
    example2 = [
        "PRJ-2024-015", "PH-001", "Concept",
        "Phase conceptuelle", "REALIZATION",
        "FORFAIT", 1,
        "2024-06-01", "2024-12-31",
        500, 75000.00, "non"
    ]
    for col, val in enumerate(example2, 1):
        ws2.cell(row=2, column=col, value=val)
    style_example(ws2, len(headers2))

    # WBS sheet
    ws3 = wb.create_sheet("WBS")
    headers3 = [
        "project_code *", "phase_ref", "parent_ref",
        "wbs_ref *", "standard_label *", "client_facing_label",
        "element_type *", "order",
        "budgeted_hours", "budgeted_cost",
        "is_billable"
    ]
    for col, h in enumerate(headers3, 1):
        ws3.cell(row=1, column=col, value=h)
    style_header(ws3, len(headers3))
    example3 = [
        "PRJ-2024-015", "PH-001", "",
        "WBS-001", "Esquisse et options conceptuelles",
        "1.2 Esquisse", "TASK", 1,
        120, 18000.00, "oui"
    ]
    for col, val in enumerate(example3, 1):
        ws3.cell(row=2, column=col, value=val)
    style_example(ws3, len(headers3))

    wb.save(os.path.join(OUTPUT_DIR, "03_projets.xlsx"))
    print("✓ 03_projets.xlsx")


def create_timesheets():
    wb = Workbook()
    add_instructions(wb, [
        "IMPORT FEUILLES DE TEMPS",
        "",
        "• Une ligne par entrée (employé + projet + phase + date + heures)",
        "• employee_username doit correspondre au fichier 01_employes.xlsx",
        "• project_code doit correspondre au fichier 03_projets.xlsx",
        "• phase_ref doit correspondre à une phase du même projet",
        "",
        "Format attendu:",
        "• date: AAAA-MM-JJ",
        "• hours: nombre décimal (ex: 7.5)",
        "• status: DRAFT, SUBMITTED, PM_APPROVED, FINANCE_APPROVED",
        "",
        "Note: Importer les 3 dernières semaines seulement pour les tests",
    ])

    ws = wb.create_sheet("Feuilles_de_temps")
    headers = [
        "employee_username *", "project_code *", "phase_ref",
        "date *", "hours *", "notes", "status"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, len(headers))
    example = [
        "jean.tremblay", "PRJ-2024-015", "PH-001",
        "2026-03-16", 7.5, "Revue esquisse concept A", "DRAFT"
    ]
    for col, val in enumerate(example, 1):
        ws.cell(row=2, column=col, value=val)
    style_example(ws, len(headers))

    wb.save(os.path.join(OUTPUT_DIR, "04_feuilles_de_temps.xlsx"))
    print("✓ 04_feuilles_de_temps.xlsx")


def create_suppliers():
    wb = Workbook()
    add_instructions(wb, [
        "IMPORT SOUS-TRAITANTS / ORGANISATIONS EXTERNES",
        "",
        "• Une ligne par organisation",
        "• type_tags: liste séparée par virgules (ex: st,partner)",
        "• Tags possibles: st, partner, competitor",
        "",
        "Format attendu:",
        "• NEQ: numéro d'entreprise du Québec (optionnel)",
    ])

    ws = wb.create_sheet("Organisations")
    headers = [
        "name *", "neq", "address", "city",
        "province", "postal_code", "country",
        "contact_name", "contact_email", "contact_phone",
        "type_tags *"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, len(headers))
    example = [
        "WSP Global", "1143859127", "1600 boul. René-Lévesque O",
        "Montréal", "Québec", "H3H 1P9", "Canada",
        "Marc Dubois", "marc.dubois@wsp.com", "514-555-0202",
        "st,partner"
    ]
    for col, val in enumerate(example, 1):
        ws.cell(row=2, column=col, value=val)
    style_example(ws, len(headers))

    wb.save(os.path.join(OUTPUT_DIR, "05_sous_traitants.xlsx"))
    print("✓ 05_sous_traitants.xlsx")


def create_ref_profils_poste():
    wb = Workbook()
    add_instructions(wb, [
        "RÉFÉRENCE — PROFILS DE POSTE",
        "",
        "• Un profil par ligne (31 archetypes dans l'architecture)",
        "• Exemples: Architecte, Urbaniste, Designer intérieur, Ingénieur MEP...",
        "• hourly_rate: taux horaire standard pour le profil (optionnel)",
    ])
    ws = wb.create_sheet("Profils")
    headers = ["name *", "description", "standard_hourly_rate"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, len(headers))
    for col, val in enumerate(["Architecte", "Architecte principal — conception", 95.00], 1):
        ws.cell(row=2, column=col, value=val)
    style_example(ws, len(headers))
    wb.save(os.path.join(OUTPUT_DIR, "R1_profils_poste.xlsx"))
    print("✓ R1_profils_poste.xlsx")


def create_ref_categories_depenses():
    wb = Workbook()
    add_instructions(wb, [
        "RÉFÉRENCE — CATÉGORIES DE DÉPENSES",
        "",
        "• Une catégorie par ligne",
        "• gl_account: code comptable GL (optionnel)",
        "• is_refacturable_default: oui/non — la dépense est-elle refacturable au client par défaut?",
        "• requires_receipt: oui/non — pièce justificative obligatoire?",
    ])
    ws = wb.create_sheet("Categories")
    headers = ["name *", "gl_account", "is_refacturable_default", "requires_receipt"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, len(headers))
    for col, val in enumerate(["Transport — Taxi/Uber", "6210", "oui", "oui"], 1):
        ws.cell(row=2, column=col, value=val)
    style_example(ws, len(headers))
    wb.save(os.path.join(OUTPUT_DIR, "R2_categories_depenses.xlsx"))
    print("✓ R2_categories_depenses.xlsx")


def create_ref_templates_projet():
    wb = Workbook()
    add_instructions(wb, [
        "RÉFÉRENCE — TEMPLATES DE PROJET",
        "",
        "• Onglet 'Templates' : un template par ligne",
        "• Onglet 'Phases_Template' : phases pré-configurées liées par template_code",
        "• contract_type: FORFAITAIRE, CONSORTIUM, CO_DEV, CONCEPTION_CONSTRUCTION",
        "• phase type: REALIZATION, SUPPORT",
        "• billing_mode: FORFAIT, HORAIRE",
    ])
    ws = wb.create_sheet("Templates")
    headers = ["template_code *", "name *", "contract_type *", "description"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, len(headers))
    for col, val in enumerate(["TPL-FORFAIT", "Forfaitaire — Standard", "FORFAITAIRE", "Phases séquentielles standard"], 1):
        ws.cell(row=2, column=col, value=val)
    style_example(ws, len(headers))

    ws2 = wb.create_sheet("Phases_Template")
    headers2 = ["template_code *", "name *", "client_label", "phase_type", "billing_mode", "is_mandatory"]
    for col, h in enumerate(headers2, 1):
        ws2.cell(row=1, column=col, value=h)
    style_header(ws2, len(headers2))
    for col, val in enumerate(["TPL-FORFAIT", "Concept", "Phase conceptuelle", "REALIZATION", "FORFAIT", "non"], 1):
        ws2.cell(row=2, column=col, value=val)
    style_example(ws2, len(headers2))
    wb.save(os.path.join(OUTPUT_DIR, "R3_templates_projet.xlsx"))
    print("✓ R3_templates_projet.xlsx")


def create_ref_templates_facture():
    wb = Workbook()
    add_instructions(wb, [
        "RÉFÉRENCE — TEMPLATES DE FACTURE",
        "",
        "• Un template par ligne",
        "• sections: liste séparée par virgules (forfait,horaire,st,depenses,retenue,taxes)",
        "• Chaque client peut utiliser un template différent",
    ])
    ws = wb.create_sheet("Templates_Facture")
    headers = ["name *", "description", "sections"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, len(headers))
    for col, val in enumerate(["Standard", "Format standard Provencher Roy", "forfait,horaire,st,depenses,retenue,taxes"], 1):
        ws.cell(row=2, column=col, value=val)
    style_example(ws, len(headers))
    wb.save(os.path.join(OUTPUT_DIR, "R4_templates_facture.xlsx"))
    print("✓ R4_templates_facture.xlsx")


def create_ref_niveaux_relance():
    wb = Workbook()
    add_instructions(wb, [
        "RÉFÉRENCE — NIVEAUX DE RELANCE (DUNNING)",
        "",
        "• 3 niveaux standard : 30 jours, 60 jours, 90 jours",
        "• email_template: texte du courriel de relance",
        "• Variables disponibles: {invoice_number}, {days}, {amount}",
    ])
    ws = wb.create_sheet("Niveaux_Relance")
    headers = ["level *", "days_overdue *", "email_template *"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, len(headers))
    for col, val in enumerate([1, 30, "Rappel courtois: la facture {invoice_number} est échue depuis {days} jours."], 1):
        ws.cell(row=2, column=col, value=val)
    style_example(ws, len(headers))
    wb.save(os.path.join(OUTPUT_DIR, "R5_niveaux_relance.xlsx"))
    print("✓ R5_niveaux_relance.xlsx")


def create_ref_unites_affaires():
    wb = Workbook()
    add_instructions(wb, [
        "RÉFÉRENCE — UNITÉS D'AFFAIRES",
        "",
        "• Une BU par ligne",
        "• Ces valeurs seront utilisées dans les projets et employés",
        "• director_username: username du directeur de BU (optionnel)",
    ])
    ws = wb.create_sheet("Unites_Affaires")
    headers = ["name *", "code", "director_username", "description"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header(ws, len(headers))
    for col, val in enumerate(["Architecture", "ARCH", "pierre.lavoie", "Unité architecture et design"], 1):
        ws.cell(row=2, column=col, value=val)
    style_example(ws, len(headers))
    wb.save(os.path.join(OUTPUT_DIR, "R6_unites_affaires.xlsx"))
    print("✓ R6_unites_affaires.xlsx")


if __name__ == "__main__":
    print("Génération des templates d'import...")
    print()
    print("=== DONNÉES DE RÉFÉRENCE ===")
    create_ref_profils_poste()
    create_ref_categories_depenses()
    create_ref_templates_projet()
    create_ref_templates_facture()
    create_ref_niveaux_relance()
    create_ref_unites_affaires()
    print()
    print("=== DONNÉES CHANGEPOINT ===")
    create_employees()
    create_clients()
    create_projects()
    create_timesheets()
    create_suppliers()
    print()
    print(f"Templates générés dans: {OUTPUT_DIR}/")
    print()
    print("Ordre d'import recommandé:")
    print("  R1-R6. Données de référence (en premier)")
    print("  1. 01_employes.xlsx")
    print("  2. 02_clients.xlsx")
    print("  3. 05_sous_traitants.xlsx")
    print("  4. 03_projets.xlsx")
    print("  5. 04_feuilles_de_temps.xlsx")
