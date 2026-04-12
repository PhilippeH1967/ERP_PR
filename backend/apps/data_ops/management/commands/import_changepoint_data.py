"""
Import ChangePoint test data from Excel files.

Usage:
    python manage.py import_changepoint_data --folder DataChagenpoint --tenant 1
    python manage.py import_changepoint_data --folder DataChagenpoint --tenant 1 --dry-run

Processes 6 files in dependency order:
  1. 01_employes.xlsx      → User + UserTenantAssociation + ProjectRole
  2. 02_clients.xlsx       → Client + Contact + ClientAddress
  3. 05_sous_traitants.xlsx → ExternalOrganization
  4. 03_projets.xlsx       → Project + Phase + Task
  5. 04_feuilles_de_temps  → TimeEntry
  6. 06_factures_fournisseurs → STInvoice

Idempotent: uses get_or_create with ChangePoint refs as dedup keys.
"""

import os
from datetime import date as dt_date
from decimal import Decimal

import openpyxl
from django.contrib.auth.models import User
from django.core.management.base import BaseCommand


def read_sheet(wb, sheet_name):
    """Read a sheet and return list of dicts (skip null rows)."""
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        return []
    headers = [str(h).replace(" *", "").replace("*", "").strip() if h else f"col_{i}" for i, h in enumerate(rows[0])]
    result = []
    for row in rows[1:]:
        if not row[0]:
            continue
        result.append({headers[i]: row[i] for i in range(min(len(headers), len(row)))})
    return result


def safe_str(val):
    """Convert value to string, handling None and numeric types."""
    if val is None:
        return ""
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    return str(val).strip()


def safe_date(val):
    """Convert value to date."""
    if val is None:
        return None
    if isinstance(val, dt_date):
        return val
    try:
        from datetime import datetime
        if isinstance(val, datetime):
            return val.date()
        return dt_date.fromisoformat(str(val)[:10])
    except (ValueError, TypeError):
        return None


def safe_decimal(val, default=0):
    """Convert value to Decimal."""
    if val is None:
        return Decimal(str(default))
    try:
        return Decimal(str(val))
    except Exception:
        return Decimal(str(default))


class Command(BaseCommand):
    help = "Import ChangePoint test data from Excel files"

    def add_arguments(self, parser):
        parser.add_argument("--folder", type=str, required=True, help="Path to DataChagenpoint folder")
        parser.add_argument("--tenant", type=int, required=True, help="Tenant ID")
        parser.add_argument("--dry-run", action="store_true", help="Simulate without writing to DB")
        parser.add_argument("--password", type=str, default="Test1234!", help="Password for created users")

    def handle(self, *args, **options):
        from apps.core.models import ProjectRole, Role, Tenant, UserTenantAssociation

        folder = options["folder"]
        tenant = Tenant.objects.get(pk=options["tenant"])
        dry_run = options["dry_run"]
        password = options["password"]

        if dry_run:
            self.stdout.write(self.style.WARNING("=== DRY RUN — aucune ecriture en base ===\n"))

        stats = {}

        # ================================================================
        # STEP 1: Employees
        # ================================================================
        self.stdout.write(self.style.MIGRATE_HEADING("Etape 1/6 : Employes"))
        wb1 = openpyxl.load_workbook(os.path.join(folder, "01_employes.xlsx"), data_only=True)
        employees = read_sheet(wb1, "Employes")
        stats["employes"] = {"created": 0, "existing": 0, "errors": 0}
        user_map = {}  # username -> User

        for emp in employees:
            username = safe_str(emp.get("username"))
            email = safe_str(emp.get("email")) or username
            first_name = safe_str(emp.get("first_name"))
            last_name = safe_str(emp.get("last_name"))
            role = safe_str(emp.get("role")) or "EMPLOYEE"
            bu = safe_str(emp.get("business_unit"))

            if not username:
                stats["employes"]["errors"] += 1
                continue

            if dry_run:
                self.stdout.write(f"  [DRY] User: {username} ({first_name} {last_name}) role={role} bu={bu}")
                user_map[username] = None
                stats["employes"]["created"] += 1
                continue

            user, created = User.objects.get_or_create(
                username=username.split("@")[0],
                defaults={"email": email, "first_name": first_name, "last_name": last_name},
            )
            if created:
                user.set_password(password)
                user.save()
                stats["employes"]["created"] += 1
            else:
                stats["employes"]["existing"] += 1

            UserTenantAssociation.objects.get_or_create(user=user, defaults={"tenant": tenant})

            # Map role string to Role enum
            role_value = role.upper()
            if hasattr(Role, role_value):
                ProjectRole.objects.get_or_create(
                    tenant=tenant, user=user, project_id=None, role=role_value,
                )
            user_map[username] = user
            user_map[email] = user  # also map by email

        self._print_stats("Employes", stats["employes"])

        # ================================================================
        # STEP 2: Clients
        # ================================================================
        self.stdout.write(self.style.MIGRATE_HEADING("Etape 2/6 : Clients"))
        wb2 = openpyxl.load_workbook(os.path.join(folder, "02_clients.xlsx"), data_only=True)
        clients_data = read_sheet(wb2, "Clients")
        contacts_data = read_sheet(wb2, "Contacts")
        addresses_data = read_sheet(wb2, "Adresses")
        stats["clients"] = {"created": 0, "existing": 0, "errors": 0}
        stats["contacts"] = {"created": 0, "existing": 0, "errors": 0}
        stats["addresses"] = {"created": 0, "existing": 0, "errors": 0}
        client_map = {}  # client_ref -> Client

        from apps.clients.models import Client, ClientAddress, Contact

        for cl in clients_data:
            ref = safe_str(cl.get("client_ref"))
            name = safe_str(cl.get("name"))
            if not ref or not name:
                stats["clients"]["errors"] += 1
                continue

            if dry_run:
                self.stdout.write(f"  [DRY] Client: {ref} — {name}")
                stats["clients"]["created"] += 1
                continue

            client, created = Client.objects.get_or_create(
                tenant=tenant, alias=ref,
                defaults={
                    "name": name,
                    "sector": safe_str(cl.get("sector")) or "Prive",
                    "status": safe_str(cl.get("status")) or "active",
                },
            )
            if created:
                stats["clients"]["created"] += 1
            else:
                stats["clients"]["existing"] += 1
            client_map[ref] = client

        # Contacts
        for ct in contacts_data:
            ref = safe_str(ct.get("client_ref"))
            name = safe_str(ct.get("name"))
            if not ref or not name or ref not in client_map:
                continue
            if dry_run:
                stats["contacts"]["created"] += 1
                continue
            _, created = Contact.objects.get_or_create(
                client=client_map[ref], name=name,
                defaults={
                    "tenant": tenant,
                    "email": safe_str(ct.get("email")),
                    "language_preference": safe_str(ct.get("language_preference")) or "fr",
                },
            )
            stats["contacts"]["created" if created else "existing"] += 1

        # Addresses
        for addr in addresses_data:
            ref = safe_str(addr.get("client_ref"))
            if not ref or ref not in client_map:
                continue
            addr_line = safe_str(addr.get("address_line_1"))
            if not addr_line:
                continue
            if dry_run:
                stats["addresses"]["created"] += 1
                continue
            _, created = ClientAddress.objects.get_or_create(
                client=client_map[ref], address_line_1=addr_line,
                defaults={
                    "tenant": tenant,
                    "address_line_2": safe_str(addr.get("address_line_2")),
                    "city": safe_str(addr.get("city")),
                    "province": safe_str(addr.get("province")) or "Quebec",
                    "postal_code": safe_str(addr.get("postal_code")),
                    "country": safe_str(addr.get("country")) or "Canada",
                    "is_billing": safe_str(addr.get("is_billing")).lower() in ("oui", "true", "1", "yes"),
                    "is_primary": safe_str(addr.get("is_primary")).lower() in ("oui", "true", "1", "yes"),
                },
            )
            stats["addresses"]["created" if created else "existing"] += 1

        self._print_stats("Clients", stats["clients"])
        self._print_stats("Contacts", stats["contacts"])
        self._print_stats("Adresses", stats["addresses"])

        # ================================================================
        # STEP 3: Subcontractors
        # ================================================================
        self.stdout.write(self.style.MIGRATE_HEADING("Etape 3/6 : Sous-traitants"))
        wb5 = openpyxl.load_workbook(os.path.join(folder, "05_sous_traitants.xlsx"), data_only=True)
        orgs_data = read_sheet(wb5, "Organisations")
        stats["orgs"] = {"created": 0, "existing": 0, "errors": 0}
        org_map = {}  # name_lower -> ExternalOrganization

        from apps.suppliers.models import ExternalOrganization

        for org in orgs_data:
            name = safe_str(org.get("name"))
            if not name:
                continue
            if dry_run:
                self.stdout.write(f"  [DRY] Org: {name}")
                stats["orgs"]["created"] += 1
                continue

            ext_org, created = ExternalOrganization.objects.get_or_create(
                tenant=tenant, name=name,
                defaults={
                    "neq": safe_str(org.get("neq")),
                    "address": safe_str(org.get("address")),
                    "city": safe_str(org.get("city")),
                    "province": safe_str(org.get("province")) or "Quebec",
                    "postal_code": safe_str(org.get("postal_code")),
                    "country": safe_str(org.get("country")) or "Canada",
                    "contact_name": safe_str(org.get("contact_name")),
                    "contact_email": safe_str(org.get("contact_email")),
                    "contact_phone": safe_str(org.get("contact_phone")),
                    "type_tags": ["st"],
                },
            )
            stats["orgs"]["created" if created else "existing"] += 1
            org_map[name.lower()] = ext_org

        self._print_stats("Organisations ST", stats["orgs"])

        # ================================================================
        # STEP 4: Projects + Phases + Tasks (WBS)
        # ================================================================
        self.stdout.write(self.style.MIGRATE_HEADING("Etape 4/6 : Projets + Phases + Taches"))
        wb3 = openpyxl.load_workbook(os.path.join(folder, "03_projets.xlsx"), data_only=True)
        projects_data = read_sheet(wb3, "Projets")
        phases_data = read_sheet(wb3, "Phases")
        wbs_data = read_sheet(wb3, "WBS")
        stats["projects"] = {"created": 0, "existing": 0, "errors": 0}
        stats["phases"] = {"created": 0, "existing": 0, "errors": 0}
        stats["tasks"] = {"created": 0, "existing": 0, "errors": 0}
        project_map = {}  # code_str -> Project
        phase_map = {}    # (project_code, phase_ref) -> Phase
        task_map = {}     # (project_code, wbs_ref) -> Task

        from apps.projects.models import Phase, Project, Task

        for proj in projects_data:
            code = safe_str(proj.get("project_code"))
            name = safe_str(proj.get("name")) or code
            if not code:
                continue

            client_ref = safe_str(proj.get("client_ref"))
            client = client_map.get(client_ref)
            pm_username = safe_str(proj.get("pm_username"))
            pm = user_map.get(pm_username)
            assoc_username = safe_str(proj.get("associe_en_charge_username"))
            assoc = user_map.get(assoc_username)
            is_internal = safe_str(proj.get("is_internal")).lower() in ("oui", "true", "1", "yes")
            contract_type = safe_str(proj.get("contract_type")) or "FORFAITAIRE"

            if dry_run:
                self.stdout.write(f"  [DRY] Project: {code} — {name} (client={client_ref}, pm={pm_username})")
                stats["projects"]["created"] += 1
                continue

            project, created = Project.objects.get_or_create(
                tenant=tenant, code=code,
                defaults={
                    "name": name,
                    "client": client,
                    "contract_type": contract_type,
                    "status": safe_str(proj.get("status")) or "ACTIVE",
                    "business_unit": safe_str(proj.get("business_unit")),
                    "start_date": safe_date(proj.get("start_date")),
                    "end_date": safe_date(proj.get("end_date")),
                    "pm": pm,
                    "associate_in_charge": assoc,
                    "is_internal": is_internal,
                },
            )
            stats["projects"]["created" if created else "existing"] += 1
            project_map[code] = project

        # Phases (from file — only project 240203 has them)
        for i, ph in enumerate(phases_data):
            proj_code = safe_str(ph.get("project_code"))
            phase_ref = safe_str(ph.get("phase_ref"))
            name = safe_str(ph.get("name"))
            project = project_map.get(proj_code)
            if not project or not phase_ref or dry_run:
                if dry_run and proj_code and phase_ref:
                    self.stdout.write(f"  [DRY] Phase: {proj_code}/{phase_ref} — {name}")
                    stats["phases"]["created"] += 1
                continue

            # Use name prefix as unique code to handle duplicates (e.g. phase_ref 0004 used twice)
            unique_code = name[:4].strip(".").strip() if name else phase_ref
            phase, created = Phase.objects.get_or_create(
                tenant=tenant, project=project, name=name,
                defaults={
                    "code": unique_code,
                    "order": i,
                    "is_mandatory": safe_str(ph.get("is_mandatory")).lower() in ("oui", "true", "1"),
                    "budgeted_hours": safe_decimal(ph.get("budgeted_hours"), 0),
                    "budgeted_cost": safe_decimal(ph.get("budgeted_cost"), 0),
                },
            )
            stats["phases"]["created" if created else "existing"] += 1
            phase_map[(proj_code, phase_ref)] = phase

        # Tasks (WBS — from file)
        for i, wbs in enumerate(wbs_data):
            proj_code = safe_str(wbs.get("project_code"))
            phase_ref = safe_str(wbs.get("phase_ref"))
            wbs_ref = safe_str(wbs.get("wbs_ref"))
            label = safe_str(wbs.get("client_facing_label")) or safe_str(wbs.get("standard_label"))
            project = project_map.get(proj_code)
            phase = phase_map.get((proj_code, phase_ref))

            if not project or not wbs_ref or dry_run:
                if dry_run and wbs_ref:
                    self.stdout.write(f"  [DRY] Task: {proj_code}/{wbs_ref} — {label}")
                    stats["tasks"]["created"] += 1
                continue

            # If phase not resolved, create a fallback
            if not phase:
                phase, _ = Phase.objects.get_or_create(
                    tenant=tenant, project=project, name="Divers",
                    defaults={"code": "DIV", "order": 99},
                )
                phase_map[(proj_code, phase_ref)] = phase

            task, created = Task.objects.get_or_create(
                tenant=tenant, project=project, wbs_code=wbs_ref[:20],
                defaults={
                    "phase": phase,
                    "name": wbs_ref[:255],
                    "client_facing_label": label[:255],
                    "order": i,
                    "budgeted_hours": safe_decimal(wbs.get("budgeted_hours"), 0),
                    "budgeted_cost": safe_decimal(wbs.get("budgeted_cost"), 0),
                    "is_billable": safe_str(wbs.get("is_billable")).lower() in ("oui", "true", "1", "yes"),
                },
            )
            stats["tasks"]["created" if created else "existing"] += 1
            task_map[(proj_code, wbs_ref)] = task

        self._print_stats("Projets", stats["projects"])
        self._print_stats("Phases", stats["phases"])
        self._print_stats("Taches WBS", stats["tasks"])

        # ================================================================
        # STEP 5: Time Entries
        # ================================================================
        self.stdout.write(self.style.MIGRATE_HEADING("Etape 5/6 : Feuilles de temps"))
        wb4 = openpyxl.load_workbook(os.path.join(folder, "04_feuilles_de_temps.xlsx"), data_only=True)
        ts_data = read_sheet(wb4, "Feuilles_de_temps")
        stats["time_entries"] = {"created": 0, "existing": 0, "errors": 0, "phases_auto_created": 0}

        from apps.time_entries.models import TimeEntry

        # Build a reverse lookup: wbs_ref -> Task (for matching timesheet phase_ref)
        wbs_ref_to_task = {}
        for (proj_code, wbs_ref), task in task_map.items():
            wbs_ref_to_task[(proj_code, wbs_ref)] = task

        for ts in ts_data:
            username = safe_str(ts.get("employee_username"))
            proj_code = safe_str(ts.get("project_code"))
            phase_ref = safe_str(ts.get("phase_ref"))
            entry_date = safe_date(ts.get("date"))
            hours = safe_decimal(ts.get("hours"), 0)
            notes = safe_str(ts.get("notes"))
            status = safe_str(ts.get("status")) or "PM_APPROVED"

            user = user_map.get(username)
            project = project_map.get(proj_code)

            if not user or not project or not entry_date:
                stats["time_entries"]["errors"] += 1
                continue

            if dry_run:
                stats["time_entries"]["created"] += 1
                continue

            # Resolve task: phase_ref in timesheets = wbs_ref in WBS
            task = wbs_ref_to_task.get((proj_code, phase_ref))
            phase = None

            if task:
                phase = task.phase
            else:
                # Auto-create phase+task for projects without WBS definition
                # Use phase_ref as both phase name and task name
                phase_key = f"auto-{proj_code}-{phase_ref[:20]}"
                if (proj_code, phase_key) not in phase_map:
                    auto_phase, ph_created = Phase.objects.get_or_create(
                        tenant=tenant, project=project, code=phase_key[:20],
                        defaults={
                            "name": phase_ref,
                            "client_facing_label": phase_ref,
                            "order": Phase.objects.filter(project=project).count(),
                        },
                    )
                    phase_map[(proj_code, phase_key)] = auto_phase
                    if ph_created:
                        stats["time_entries"]["phases_auto_created"] += 1

                phase = phase_map[(proj_code, phase_key)]

                # Auto-create task under that phase
                if (proj_code, phase_ref) not in task_map:
                    auto_task, _ = Task.objects.get_or_create(
                        tenant=tenant, project=project, wbs_code=phase_ref[:20],
                        defaults={
                            "phase": phase,
                            "name": phase_ref,
                            "client_facing_label": phase_ref,
                            "order": Task.objects.filter(project=project, phase=phase).count(),
                        },
                    )
                    task_map[(proj_code, phase_ref)] = auto_task
                    wbs_ref_to_task[(proj_code, phase_ref)] = auto_task

                task = wbs_ref_to_task.get((proj_code, phase_ref))

            _, created = TimeEntry.objects.get_or_create(
                tenant=tenant,
                employee=user,
                project=project,
                task=task,
                date=entry_date,
                defaults={
                    "phase": phase,
                    "hours": hours,
                    "notes": notes,
                    "status": status,
                },
            )
            stats["time_entries"]["created" if created else "existing"] += 1

        self._print_stats("Feuilles de temps", stats["time_entries"])
        if stats["time_entries"]["phases_auto_created"]:
            self.stdout.write(
                self.style.WARNING(
                    f"    Phases/taches auto-creees: {stats['time_entries']['phases_auto_created']}"
                )
            )

        # ================================================================
        # STEP 6: Supplier Invoices
        # ================================================================
        self.stdout.write(self.style.MIGRATE_HEADING("Etape 6/6 : Factures fournisseurs"))
        wb6 = openpyxl.load_workbook(os.path.join(folder, "06_factures_fournisseurs.xlsx"), data_only=True)
        inv_data = read_sheet(wb6, "Factures fournisseurs")
        stats["st_invoices"] = {"created": 0, "existing": 0, "errors": 0}

        from apps.suppliers.models import STInvoice

        for inv in inv_data:
            inv_number = safe_str(inv.get("no_facture"))
            supplier_name = safe_str(inv.get("nom_fournisseur"))
            proj_code = safe_str(inv.get("code_projet"))
            amount = safe_decimal(inv.get("montant"), 0)
            inv_date = safe_date(inv.get("date_facture"))
            status = safe_str(inv.get("statut")) or "received"
            description = safe_str(inv.get("description"))
            cat = safe_str(inv.get("categorie_budget")).lower()

            project = project_map.get(proj_code)

            # Fuzzy match supplier
            supplier = None
            supplier_lower = supplier_name.lower()
            for org_name, org in org_map.items():
                if org_name in supplier_lower or supplier_lower in org_name:
                    supplier = org
                    break

            if not project or not supplier or not inv_date:
                self.stdout.write(self.style.WARNING(
                    f"  SKIP: facture {inv_number} — projet={proj_code} fournisseur={supplier_name} (non resolu)"
                ))
                stats["st_invoices"]["errors"] += 1
                continue

            if dry_run:
                self.stdout.write(f"  [DRY] STInvoice: {inv_number} — {supplier_name} — {amount}$")
                stats["st_invoices"]["created"] += 1
                continue

            budget_fields = {"budget_internal": 0, "budget_refacturable": 0, "budget_absorbed": 0}
            if "honor" in cat:
                budget_fields["budget_internal"] = amount
            elif "refact" in cat:
                budget_fields["budget_refacturable"] = amount
            elif "absorb" in cat:
                budget_fields["budget_absorbed"] = amount

            _, created = STInvoice.objects.get_or_create(
                tenant=tenant, invoice_number=inv_number,
                defaults={
                    "project": project,
                    "supplier": supplier,
                    "invoice_date": inv_date,
                    "amount": amount,
                    "status": status,
                    **budget_fields,
                },
            )
            stats["st_invoices"]["created" if created else "existing"] += 1

        self._print_stats("Factures ST", stats["st_invoices"])

        # ================================================================
        # RECONCILIATION REPORT
        # ================================================================
        self.stdout.write("\n" + "=" * 60)
        self.stdout.write(self.style.SUCCESS("RAPPORT DE RECONCILIATION"))
        self.stdout.write("=" * 60)

        total_created = 0
        total_existing = 0
        total_errors = 0

        for label, s in stats.items():
            c = s.get("created", 0)
            e = s.get("existing", 0)
            err = s.get("errors", 0)
            total_created += c
            total_existing += e
            total_errors += err
            icon = "+" if c > 0 else "="
            self.stdout.write(f"  {icon} {label:25s}  crees={c:>4}  existants={e:>4}  erreurs={err:>4}")

        self.stdout.write("-" * 60)
        self.stdout.write(f"  TOTAL                     crees={total_created:>4}  existants={total_existing:>4}  erreurs={total_errors:>4}")

        if dry_run:
            self.stdout.write(self.style.WARNING("\n=== DRY RUN — rien n'a ete ecrit en base ==="))
        else:
            self.stdout.write(self.style.SUCCESS(f"\nImport termine avec succes! {total_created} objets crees."))

    def _print_stats(self, label, s):
        c = s.get("created", 0)
        e = s.get("existing", 0)
        err = s.get("errors", 0)
        style = self.style.SUCCESS if err == 0 else self.style.WARNING
        self.stdout.write(style(f"  {label}: {c} crees, {e} existants, {err} erreurs"))
