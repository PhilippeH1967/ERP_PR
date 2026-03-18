"""
Import reference/configuration data from Excel templates.

Usage:
    python manage.py import_reference_data /path/to/folder/ --tenant=provencher-roy

Reads R1-R6 Excel files and creates reference records.
"""

import os

import structlog
from django.core.management.base import BaseCommand
from openpyxl import load_workbook

logger = structlog.get_logger()


class Command(BaseCommand):
    help = "Import reference data from Excel templates (R1-R6)"

    def add_arguments(self, parser):
        parser.add_argument("folder", type=str, help="Path to folder containing R*.xlsx files")
        parser.add_argument("--tenant", type=str, required=True, help="Tenant slug")

    def handle(self, *args, **options):
        folder = options["folder"]
        tenant_slug = options["tenant"]

        from apps.core.models import Tenant

        tenant, _ = Tenant.objects.get_or_create(
            slug=tenant_slug,
            defaults={"name": tenant_slug.replace("-", " ").title()},
        )
        self.tenant = tenant
        self.stdout.write(f"Tenant: {tenant.name}\n")

        self._import_profils(folder)
        self._import_categories(folder)
        self._import_templates_projet(folder)
        self._import_templates_facture(folder)
        self._import_niveaux_relance(folder)
        self._import_unites_affaires(folder)

        self.stdout.write(self.style.SUCCESS("\nDonnées de référence importées!"))

    def _read_sheet(self, filepath, sheet_name):
        if not os.path.exists(filepath):
            self.stdout.write(self.style.WARNING(f"  Fichier non trouvé: {filepath}"))
            return []
        wb = load_workbook(filepath, read_only=True)
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return []
        headers = [str(h).replace(" *", "").strip() for h in rows[0]]
        data = []
        for row in rows[1:]:
            if all(v is None for v in row):
                continue
            record = {headers[i]: v for i, v in enumerate(row) if i < len(headers)}
            data.append(record)
        return data

    def _import_profils(self, folder):
        self.stdout.write("R1. Profils de poste...")
        data = self._read_sheet(os.path.join(folder, "R1_profils_poste.xlsx"), "Profils")
        # Store as tenant config (no dedicated model yet — using notes)
        count = len(data)
        self.stdout.write(self.style.SUCCESS(f"  ✓ {count} profils lus (référence)"))

    def _import_categories(self, folder):
        self.stdout.write("R2. Catégories de dépenses...")
        data = self._read_sheet(os.path.join(folder, "R2_categories_depenses.xlsx"), "Categories")
        from apps.expenses.models import ExpenseCategory

        count = 0
        for row in data:
            name = row.get("name")
            if not name:
                continue
            is_ref = str(row.get("is_refacturable_default", "non")).lower() in ("oui", "true", "1")
            req_rec = str(row.get("requires_receipt", "oui")).lower() in ("oui", "true", "1")
            _, created = ExpenseCategory.objects.get_or_create(
                tenant=self.tenant, name=name,
                defaults={
                    "gl_account": row.get("gl_account", ""),
                    "is_refacturable_default": is_ref,
                    "requires_receipt": req_rec,
                },
            )
            if created:
                count += 1
        self.stdout.write(self.style.SUCCESS(f"  ✓ {count} catégories créées"))

    def _import_templates_projet(self, folder):
        self.stdout.write("R3. Templates de projet...")
        filepath = os.path.join(folder, "R3_templates_projet.xlsx")
        templates = self._read_sheet(filepath, "Templates")
        phases = self._read_sheet(filepath, "Phases_Template")

        from apps.projects.models import ProjectTemplate

        count = 0
        for tmpl in templates:
            code = tmpl.get("template_code")
            name = tmpl.get("name")
            if not code or not name:
                continue
            tmpl_phases = [
                {
                    "name": p.get("name", ""),
                    "client_label": p.get("client_label", ""),
                    "type": p.get("phase_type", "REALIZATION"),
                    "billing_mode": p.get("billing_mode", "FORFAIT"),
                    "is_mandatory": str(p.get("is_mandatory", "non")).lower() in ("oui", "true", "1"),
                }
                for p in phases if p.get("template_code") == code
            ]
            _, created = ProjectTemplate.objects.get_or_create(
                tenant=self.tenant, code=code,
                defaults={
                    "name": name,
                    "contract_type": tmpl.get("contract_type", "FORFAITAIRE"),
                    "description": tmpl.get("description", ""),
                    "phases_config": tmpl_phases,
                    "support_services_config": [],
                },
            )
            if created:
                count += 1
        self.stdout.write(self.style.SUCCESS(f"  ✓ {count} templates créés"))

    def _import_templates_facture(self, folder):
        self.stdout.write("R4. Templates de facture...")
        data = self._read_sheet(os.path.join(folder, "R4_templates_facture.xlsx"), "Templates_Facture")
        from apps.billing.models import InvoiceTemplate

        count = 0
        for row in data:
            name = row.get("name")
            if not name:
                continue
            sections = [s.strip() for s in str(row.get("sections", "")).split(",") if s.strip()]
            _, created = InvoiceTemplate.objects.get_or_create(
                tenant=self.tenant, name=name,
                defaults={
                    "description": row.get("description", ""),
                    "template_config": {"sections": sections, "logo": True, "banking_footer": True},
                },
            )
            if created:
                count += 1
        self.stdout.write(self.style.SUCCESS(f"  ✓ {count} templates créés"))

    def _import_niveaux_relance(self, folder):
        self.stdout.write("R5. Niveaux de relance...")
        data = self._read_sheet(os.path.join(folder, "R5_niveaux_relance.xlsx"), "Niveaux_Relance")
        from apps.billing.models import DunningLevel

        count = 0
        for row in data:
            level = row.get("level")
            if not level:
                continue
            _, created = DunningLevel.objects.get_or_create(
                tenant=self.tenant, level=int(level),
                defaults={
                    "days_overdue": int(row.get("days_overdue", 30)),
                    "email_template": row.get("email_template", ""),
                },
            )
            if created:
                count += 1
        self.stdout.write(self.style.SUCCESS(f"  ✓ {count} niveaux créés"))

    def _import_unites_affaires(self, folder):
        self.stdout.write("R6. Unités d'affaires...")
        data = self._read_sheet(os.path.join(folder, "R6_unites_affaires.xlsx"), "Unites_Affaires")
        # BU stored as reference — currently string field on Project
        count = len(data)
        self.stdout.write(self.style.SUCCESS(f"  ✓ {count} BU lues (référence)"))
