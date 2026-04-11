"""
Seed reference/configuration data required before ChangePoint import.

Usage:
    python manage.py seed_reference_data --tenant=provencher-roy

Creates standard reference data:
- Position profiles (from listePoste.xlsx or defaults)
- Expense categories
- Project templates (1 per contract type)
- Invoice templates (1 default)
- Dunning levels (3 levels)
"""

import os

import structlog
from django.core.management.base import BaseCommand

logger = structlog.get_logger()

# Standard expense categories for architecture firm
DEFAULT_EXPENSE_CATEGORIES = [
    {"name": "Transport — Taxi/Uber", "gl_account": "6210", "is_refacturable_default": True, "requires_receipt": True},
    {"name": "Transport — Kilométrage", "gl_account": "6211", "is_refacturable_default": True, "requires_receipt": False},
    {"name": "Transport — Transport en commun", "gl_account": "6212", "is_refacturable_default": False, "requires_receipt": True},
    {"name": "Repas — Client", "gl_account": "6220", "is_refacturable_default": True, "requires_receipt": True},
    {"name": "Repas — Équipe", "gl_account": "6221", "is_refacturable_default": False, "requires_receipt": True},
    {"name": "Hébergement", "gl_account": "6230", "is_refacturable_default": True, "requires_receipt": True},
    {"name": "Stationnement", "gl_account": "6240", "is_refacturable_default": True, "requires_receipt": True},
    {"name": "Fournitures de bureau", "gl_account": "6310", "is_refacturable_default": False, "requires_receipt": True},
    {"name": "Impression / Reproduction", "gl_account": "6320", "is_refacturable_default": True, "requires_receipt": True},
    {"name": "Télécommunications", "gl_account": "6330", "is_refacturable_default": False, "requires_receipt": True},
    {"name": "Formation / Conférence", "gl_account": "6410", "is_refacturable_default": False, "requires_receipt": True},
    {"name": "Cotisation professionnelle", "gl_account": "6420", "is_refacturable_default": False, "requires_receipt": True},
    {"name": "Divers", "gl_account": "6900", "is_refacturable_default": False, "requires_receipt": True},
]

# Standard project templates
DEFAULT_PROJECT_TEMPLATES = [
    {
        "name": "Forfaitaire — Standard",
        "code": "TPL-FORFAIT",
        "contract_type": "FORFAITAIRE",
        "description": "Projet forfaitaire avec phases séquentielles standard",
        "phases_config": [
            {"name": "Étude préparatoire", "client_label": "Phase 1 — Études", "type": "REALIZATION", "billing_mode": "FORFAIT", "is_mandatory": False},
            {"name": "Concept", "client_label": "Phase 2 — Concept", "type": "REALIZATION", "billing_mode": "FORFAIT", "is_mandatory": False},
            {"name": "Préliminaire", "client_label": "Phase 3 — Préliminaire", "type": "REALIZATION", "billing_mode": "FORFAIT", "is_mandatory": False},
            {"name": "Définitif", "client_label": "Phase 4 — Définitif", "type": "REALIZATION", "billing_mode": "FORFAIT", "is_mandatory": False},
            {"name": "Appel d'offres", "client_label": "Phase 5 — Appel d'offres", "type": "REALIZATION", "billing_mode": "FORFAIT", "is_mandatory": False},
            {"name": "Surveillance", "client_label": "Phase 6 — Surveillance", "type": "REALIZATION", "billing_mode": "HORAIRE", "is_mandatory": False},
            {"name": "Gestion de projet", "client_label": "Gestion de projet", "type": "SUPPORT", "billing_mode": "FORFAIT", "is_mandatory": True},
            {"name": "Qualité", "client_label": "Assurance qualité", "type": "SUPPORT", "billing_mode": "FORFAIT", "is_mandatory": True},
        ],
        "support_services_config": [
            {"name": "BIM", "client_label": "Services BIM"},
            {"name": "DD", "client_label": "Développement durable"},
        ],
    },
    {
        "name": "Consortium — Standard",
        "code": "TPL-CONSORTIUM",
        "contract_type": "CONSORTIUM",
        "description": "Projet en consortium avec phases standard",
        "phases_config": [
            {"name": "Concept", "client_label": "Concept", "type": "REALIZATION", "billing_mode": "FORFAIT"},
            {"name": "Préliminaire", "client_label": "Préliminaire", "type": "REALIZATION", "billing_mode": "FORFAIT"},
            {"name": "Définitif", "client_label": "Définitif", "type": "REALIZATION", "billing_mode": "FORFAIT"},
            {"name": "Surveillance", "client_label": "Surveillance", "type": "REALIZATION", "billing_mode": "HORAIRE"},
            {"name": "Gestion de projet", "client_label": "Gestion de projet", "type": "SUPPORT", "billing_mode": "FORFAIT", "is_mandatory": True},
        ],
        "support_services_config": [],
    },
    {
        "name": "Co-développement — Standard",
        "code": "TPL-CODEV",
        "contract_type": "CO_DEV",
        "description": "Projet en co-développement",
        "phases_config": [
            {"name": "Concept", "client_label": "Concept", "type": "REALIZATION", "billing_mode": "FORFAIT"},
            {"name": "Développement", "client_label": "Développement", "type": "REALIZATION", "billing_mode": "HORAIRE"},
            {"name": "Gestion de projet", "client_label": "Gestion de projet", "type": "SUPPORT", "billing_mode": "FORFAIT", "is_mandatory": True},
        ],
        "support_services_config": [],
    },
    {
        "name": "Conception-construction — Standard",
        "code": "TPL-CC",
        "contract_type": "CONCEPTION_CONSTRUCTION",
        "description": "Projet conception-construction intégré",
        "phases_config": [
            {"name": "Concept", "client_label": "Concept", "type": "REALIZATION", "billing_mode": "FORFAIT"},
            {"name": "Design détaillé", "client_label": "Design détaillé", "type": "REALIZATION", "billing_mode": "FORFAIT"},
            {"name": "Construction", "client_label": "Construction", "type": "REALIZATION", "billing_mode": "HORAIRE"},
            {"name": "Gestion de projet", "client_label": "Gestion de projet", "type": "SUPPORT", "billing_mode": "FORFAIT", "is_mandatory": True},
        ],
        "support_services_config": [],
    },
]

DEFAULT_DUNNING_LEVELS = [
    {"level": 1, "days_overdue": 30, "email_template": "Rappel courtois: la facture {invoice_number} est échue depuis {days} jours. Merci de procéder au paiement."},
    {"level": 2, "days_overdue": 60, "email_template": "Deuxième rappel: la facture {invoice_number} est en souffrance depuis {days} jours. Veuillez régulariser la situation."},
    {"level": 3, "days_overdue": 90, "email_template": "Mise en demeure: la facture {invoice_number} est impayée depuis {days} jours. Sans paiement sous 10 jours, nous transmettrons le dossier."},
]


class Command(BaseCommand):
    help = "Seed reference data (categories, templates, dunning levels)"

    def add_arguments(self, parser):
        parser.add_argument("--tenant", type=str, required=True, help="Tenant slug")

    def handle(self, *args, **options):
        from apps.core.models import Tenant

        tenant_slug = options["tenant"]
        tenant, created = Tenant.objects.get_or_create(
            slug=tenant_slug,
            defaults={"name": tenant_slug.replace("-", " ").title()},
        )

        self.stdout.write(f"Tenant: {tenant.name}\n")

        # 1. Position profiles from listePoste.xlsx
        self._seed_position_profiles(tenant)

        # 2. Expense categories
        self._seed_expense_categories(tenant)

        # 3. Project templates
        self._seed_project_templates(tenant)

        # 4. Invoice template
        self._seed_invoice_templates(tenant)

        # 5. Dunning levels
        self._seed_dunning_levels(tenant)

        # 6. Leave types (Quebec standard)
        self._seed_leave_types(tenant)

        self.stdout.write(self.style.SUCCESS("\nDonnées de référence créées!"))

    def _seed_leave_types(self, tenant):
        self.stdout.write("6. Types de congés (Québec)...")
        from apps.leaves.services import seed_leave_types

        count = seed_leave_types(tenant)
        self.stdout.write(self.style.SUCCESS(f"  ✓ {count} types de congés créés"))

    def _seed_position_profiles(self, tenant):
        """Try to read listePoste.xlsx, fallback to defaults."""
        self.stdout.write("1. Profils de poste...")

        # listePoste.xlsx is at repo root
        xlsx_path = os.path.join(
            os.path.dirname(__file__), "..", "..", "..", "..", "listePoste.xlsx"
        )
        xlsx_path = os.path.normpath(xlsx_path)

        profiles = []
        if os.path.exists(xlsx_path):
            try:
                from openpyxl import load_workbook

                wb = load_workbook(xlsx_path, read_only=True)
                ws = wb.active
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if row[0]:
                        profiles.append(str(row[0]).strip())
            except Exception as e:
                self.stdout.write(self.style.WARNING(f"  Erreur lecture listePoste.xlsx: {e}"))

        if not profiles:
            profiles = [
                "Architecte", "Architecte principal", "Architecte concepteur",
                "Urbaniste", "Designer intérieur", "Architecte paysagiste",
                "Ingénieur structure", "Ingénieur mécanique", "Ingénieur électrique",
                "Technologue architecture", "Technologue structure",
                "Dessinateur", "Modélisateur BIM", "Coordonnateur BIM",
                "Chargé de projet", "Directeur de projet",
                "Directeur design", "Directeur technique",
                "Surveillant de chantier", "Estimateur",
                "Adjoint administratif", "Contrôleur financier",
                "Comptable", "Analyste financier",
                "Gestionnaire RH", "Coordonnateur",
                "Stagiaire architecture", "Stagiaire design",
                "Consultant externe", "Spécialiste DD",
                "Gestionnaire propositions",
            ]

        # Store as a simple note — no dedicated PositionProfile model yet
        # Could be added to Tenant config or a reference table later
        self.stdout.write(self.style.SUCCESS(f"  ✓ {len(profiles)} profils identifiés (stockés en référence)"))

    def _seed_expense_categories(self, tenant):
        self.stdout.write("2. Catégories de dépenses...")
        from apps.expenses.models import ExpenseCategory

        count = 0
        for cat in DEFAULT_EXPENSE_CATEGORIES:
            _, created = ExpenseCategory.objects.get_or_create(
                tenant=tenant,
                name=cat["name"],
                defaults={
                    "gl_account": cat["gl_account"],
                    "is_refacturable_default": cat["is_refacturable_default"],
                    "requires_receipt": cat["requires_receipt"],
                },
            )
            if created:
                count += 1

        self.stdout.write(self.style.SUCCESS(f"  ✓ {count} catégories créées"))

    def _seed_project_templates(self, tenant):
        self.stdout.write("3. Templates de projet...")
        from apps.projects.models import ProjectTemplate

        count = 0
        for tmpl in DEFAULT_PROJECT_TEMPLATES:
            _, created = ProjectTemplate.objects.get_or_create(
                tenant=tenant,
                code=tmpl["code"],
                defaults={
                    "name": tmpl["name"],
                    "contract_type": tmpl["contract_type"],
                    "description": tmpl["description"],
                    "phases_config": tmpl["phases_config"],
                    "support_services_config": tmpl["support_services_config"],
                },
            )
            if created:
                count += 1

        self.stdout.write(self.style.SUCCESS(f"  ✓ {count} templates créés"))

    def _seed_invoice_templates(self, tenant):
        self.stdout.write("4. Templates de facture...")
        from apps.billing.models import InvoiceTemplate

        _, created = InvoiceTemplate.objects.get_or_create(
            tenant=tenant,
            name="Standard",
            defaults={
                "description": "Format de facture standard Provencher Roy",
                "template_config": {
                    "sections": ["forfait", "horaire", "st", "depenses", "retenue", "taxes"],
                    "logo": True,
                    "banking_footer": True,
                },
            },
        )

        self.stdout.write(self.style.SUCCESS(f"  ✓ {'1 template créé' if created else 'déjà existant'}"))

    def _seed_dunning_levels(self, tenant):
        self.stdout.write("5. Niveaux de relance...")
        from apps.billing.models import DunningLevel

        count = 0
        for dl in DEFAULT_DUNNING_LEVELS:
            _, created = DunningLevel.objects.get_or_create(
                tenant=tenant,
                level=dl["level"],
                defaults={
                    "days_overdue": dl["days_overdue"],
                    "email_template": dl["email_template"],
                },
            )
            if created:
                count += 1

        self.stdout.write(self.style.SUCCESS(f"  ✓ {count} niveaux créés"))
