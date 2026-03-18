"""
Import ChangePoint data from Excel templates.

Usage:
    python manage.py import_changepoint /path/to/folder/ --tenant=provencher-roy

The folder must contain:
    01_employes.xlsx
    02_clients.xlsx
    03_projets.xlsx
    04_feuilles_de_temps.xlsx
    05_sous_traitants.xlsx

Import order: employees → clients → suppliers → projects → timesheets
"""

import os
from datetime import date

import structlog
from django.contrib.auth.models import User
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from openpyxl import load_workbook

logger = structlog.get_logger()


class Command(BaseCommand):
    help = "Import ChangePoint data from Excel templates"

    def add_arguments(self, parser):
        parser.add_argument("folder", type=str, help="Path to folder containing Excel files")
        parser.add_argument("--tenant", type=str, required=True, help="Tenant slug")
        parser.add_argument("--dry-run", action="store_true", help="Validate only, don't import")

    def handle(self, *args, **options):
        folder = options["folder"]
        tenant_slug = options["tenant"]
        dry_run = options["dry_run"]

        if not os.path.isdir(folder):
            raise CommandError(f"Folder not found: {folder}")

        from apps.core.models import Tenant

        tenant, created = Tenant.objects.get_or_create(
            slug=tenant_slug,
            defaults={"name": tenant_slug.replace("-", " ").title(), "is_active": True},
        )
        if created:
            self.stdout.write(self.style.SUCCESS(f"Tenant created: {tenant.name}"))
        else:
            self.stdout.write(f"Using tenant: {tenant.name}")

        self.tenant = tenant
        self.dry_run = dry_run
        self.refs = {"users": {}, "clients": {}, "phases": {}}

        try:
            with transaction.atomic():
                self._import_employees(folder)
                self._import_clients(folder)
                self._import_suppliers(folder)
                self._import_projects(folder)
                self._import_timesheets(folder)

                if dry_run:
                    self.stdout.write(self.style.WARNING("DRY RUN — rolling back"))
                    raise DryRunRollback()

        except DryRunRollback:
            pass

        self.stdout.write(self.style.SUCCESS("\nImport terminé!"))

    def _read_sheet(self, filepath, sheet_name):
        """Read Excel sheet and return list of dicts."""
        if not os.path.exists(filepath):
            self.stdout.write(self.style.WARNING(f"  Fichier non trouvé: {filepath}"))
            return []

        wb = load_workbook(filepath, read_only=True)
        if sheet_name not in wb.sheetnames:
            self.stdout.write(self.style.WARNING(f"  Onglet '{sheet_name}' non trouvé dans {filepath}"))
            return []

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            return []

        headers = [str(h).replace(" *", "").strip() for h in rows[0]]
        data = []
        for row in rows[1:]:
            if all(v is None for v in row):
                continue
            record = {}
            for i, val in enumerate(row):
                if i < len(headers):
                    record[headers[i]] = val
            data.append(record)
        return data

    def _import_employees(self, folder):
        self.stdout.write("\n1. Import employés...")
        data = self._read_sheet(os.path.join(folder, "01_employes.xlsx"), "Employes")
        count = 0
        for row in data:
            username = row.get("username")
            if not username:
                continue
            user, created = User.objects.get_or_create(
                username=username,
                defaults={
                    "email": row.get("email", ""),
                    "first_name": row.get("first_name", ""),
                    "last_name": row.get("last_name", ""),
                },
            )
            if created:
                user.set_unusable_password()
                user.save()

            # Create tenant association
            from apps.core.models import UserTenantAssociation

            UserTenantAssociation.objects.get_or_create(
                user=user, defaults={"tenant": self.tenant}
            )

            # Create project role if specified
            role = row.get("role", "EMPLOYEE")
            if role:
                from apps.core.models import ProjectRole

                ProjectRole.objects.get_or_create(
                    tenant=self.tenant, user=user, project_id=None, role=role,
                )

            self.refs["users"][username] = user
            count += 1

        self.stdout.write(self.style.SUCCESS(f"  ✓ {count} employés importés"))

    def _import_clients(self, folder):
        self.stdout.write("\n2. Import clients...")
        filepath = os.path.join(folder, "02_clients.xlsx")

        # Clients
        data = self._read_sheet(filepath, "Clients")
        count = 0
        for row in data:
            ref = row.get("client_ref")
            name = row.get("name")
            if not ref or not name:
                continue

            from apps.clients.models import Client

            client, _ = Client.objects.get_or_create(
                tenant=self.tenant,
                name=name,
                defaults={
                    "alias": row.get("alias", ""),
                    "legal_entity": row.get("legal_entity", ""),
                    "sector": row.get("sector", ""),
                    "status": row.get("status", "active"),
                    "payment_terms_days": int(row.get("payment_terms_days", 30) or 30),
                    "associe_en_charge": row.get("associe_en_charge", ""),
                    "notes": row.get("notes", ""),
                },
            )
            self.refs["clients"][ref] = client
            count += 1

        self.stdout.write(self.style.SUCCESS(f"  ✓ {count} clients importés"))

        # Contacts
        contacts_data = self._read_sheet(filepath, "Contacts")
        c_count = 0
        for row in contacts_data:
            ref = row.get("client_ref")
            client = self.refs["clients"].get(ref)
            if not client:
                continue

            from apps.clients.models import Contact

            Contact.objects.get_or_create(
                tenant=self.tenant,
                client=client,
                name=row.get("name", ""),
                defaults={
                    "role": row.get("role", ""),
                    "email": row.get("email", ""),
                    "phone": row.get("phone", ""),
                    "language_preference": row.get("language_preference", "fr"),
                },
            )
            c_count += 1

        self.stdout.write(self.style.SUCCESS(f"  ✓ {c_count} contacts importés"))

        # Addresses
        addr_data = self._read_sheet(filepath, "Adresses")
        a_count = 0
        for row in addr_data:
            ref = row.get("client_ref")
            client = self.refs["clients"].get(ref)
            if not client:
                continue

            from apps.clients.models import ClientAddress

            is_billing = str(row.get("is_billing", "")).lower() in ("oui", "true", "1", "yes")
            is_primary = str(row.get("is_primary", "")).lower() in ("oui", "true", "1", "yes")

            ClientAddress.objects.get_or_create(
                tenant=self.tenant,
                client=client,
                address_line_1=row.get("address_line_1", ""),
                defaults={
                    "address_line_2": row.get("address_line_2", ""),
                    "city": row.get("city", ""),
                    "province": row.get("province", "Québec"),
                    "postal_code": row.get("postal_code", ""),
                    "country": row.get("country", "Canada"),
                    "is_billing": is_billing,
                    "is_primary": is_primary,
                },
            )
            a_count += 1

        self.stdout.write(self.style.SUCCESS(f"  ✓ {a_count} adresses importées"))

    def _import_suppliers(self, folder):
        self.stdout.write("\n3. Import sous-traitants...")
        data = self._read_sheet(
            os.path.join(folder, "05_sous_traitants.xlsx"), "Organisations"
        )
        count = 0
        for row in data:
            name = row.get("name")
            if not name:
                continue

            from apps.suppliers.models import ExternalOrganization

            tags = [t.strip() for t in str(row.get("type_tags", "st")).split(",") if t.strip()]

            ExternalOrganization.objects.get_or_create(
                tenant=self.tenant,
                name=name,
                defaults={
                    "neq": row.get("neq", ""),
                    "address": row.get("address", ""),
                    "city": row.get("city", ""),
                    "province": row.get("province", "Québec"),
                    "postal_code": row.get("postal_code", ""),
                    "country": row.get("country", "Canada"),
                    "contact_name": row.get("contact_name", ""),
                    "contact_email": row.get("contact_email", ""),
                    "contact_phone": row.get("contact_phone", ""),
                    "type_tags": tags,
                },
            )
            count += 1

        self.stdout.write(self.style.SUCCESS(f"  ✓ {count} organisations importées"))

    def _import_projects(self, folder):
        self.stdout.write("\n4. Import projets...")
        filepath = os.path.join(folder, "03_projets.xlsx")

        # Projects
        data = self._read_sheet(filepath, "Projets")
        count = 0
        for row in data:
            code = row.get("project_code")
            name = row.get("name")
            if not code or not name:
                continue

            client_ref = row.get("client_ref")
            client = self.refs["clients"].get(client_ref)

            pm_username = row.get("pm_username")
            pm = self.refs["users"].get(pm_username)

            assoc_username = row.get("associe_en_charge_username")
            assoc = self.refs["users"].get(assoc_username)

            is_internal = str(row.get("is_internal", "")).lower() in ("oui", "true", "1", "yes")

            from apps.projects.models import Project

            project, _ = Project.objects.get_or_create(
                tenant=self.tenant,
                code=code,
                defaults={
                    "name": name,
                    "client": client,
                    "contract_type": row.get("contract_type", "FORFAITAIRE"),
                    "status": row.get("status", "ACTIVE"),
                    "business_unit": row.get("business_unit", ""),
                    "start_date": row.get("start_date"),
                    "end_date": row.get("end_date"),
                    "pm": pm,
                    "associate_in_charge": assoc,
                    "is_internal": is_internal,
                },
            )
            count += 1

            # Create PM project role if PM specified
            if pm:
                from apps.core.models import ProjectRole

                ProjectRole.objects.get_or_create(
                    tenant=self.tenant, user=pm, project_id=project.pk, role="PM",
                )

        self.stdout.write(self.style.SUCCESS(f"  ✓ {count} projets importés"))

        # Phases
        phases_data = self._read_sheet(filepath, "Phases")
        ph_count = 0
        for row in phases_data:
            project_code = row.get("project_code")
            phase_ref = row.get("phase_ref")
            if not project_code or not phase_ref:
                continue

            from apps.projects.models import Phase, Project

            try:
                project = Project.objects.get(tenant=self.tenant, code=project_code)
            except Project.DoesNotExist:
                continue

            is_mandatory = str(row.get("is_mandatory", "")).lower() in ("oui", "true", "1", "yes")

            phase, _ = Phase.objects.get_or_create(
                tenant=self.tenant,
                project=project,
                name=row.get("name", phase_ref),
                defaults={
                    "code": phase_ref,
                    "client_facing_label": row.get("client_facing_label", ""),
                    "phase_type": row.get("phase_type", "REALIZATION"),
                    "billing_mode": row.get("billing_mode", "FORFAIT"),
                    "order": int(row.get("order", 0) or 0),
                    "start_date": row.get("start_date"),
                    "end_date": row.get("end_date"),
                    "budgeted_hours": float(row.get("budgeted_hours", 0) or 0),
                    "budgeted_cost": float(row.get("budgeted_cost", 0) or 0),
                    "is_mandatory": is_mandatory,
                },
            )
            self.refs["phases"][f"{project_code}:{phase_ref}"] = phase
            ph_count += 1

        self.stdout.write(self.style.SUCCESS(f"  ✓ {ph_count} phases importées"))

        # WBS
        wbs_data = self._read_sheet(filepath, "WBS")
        wbs_count = 0
        for row in wbs_data:
            project_code = row.get("project_code")
            wbs_ref = row.get("wbs_ref")
            if not project_code or not wbs_ref:
                continue

            from apps.projects.models import Project, WBSElement

            try:
                project = Project.objects.get(tenant=self.tenant, code=project_code)
            except Project.DoesNotExist:
                continue

            phase_ref = row.get("phase_ref")
            phase = self.refs["phases"].get(f"{project_code}:{phase_ref}")

            is_billable = str(row.get("is_billable", "oui")).lower() in ("oui", "true", "1", "yes")

            WBSElement.objects.get_or_create(
                tenant=self.tenant,
                project=project,
                standard_label=row.get("standard_label", wbs_ref),
                defaults={
                    "phase": phase,
                    "client_facing_label": row.get("client_facing_label", ""),
                    "element_type": row.get("element_type", "TASK"),
                    "order": int(row.get("order", 0) or 0),
                    "budgeted_hours": float(row.get("budgeted_hours", 0) or 0),
                    "budgeted_cost": float(row.get("budgeted_cost", 0) or 0),
                    "is_billable": is_billable,
                },
            )
            wbs_count += 1

        self.stdout.write(self.style.SUCCESS(f"  ✓ {wbs_count} éléments WBS importés"))

    def _import_timesheets(self, folder):
        self.stdout.write("\n5. Import feuilles de temps...")
        data = self._read_sheet(
            os.path.join(folder, "04_feuilles_de_temps.xlsx"), "Feuilles_de_temps"
        )
        count = 0
        for row in data:
            username = row.get("employee_username")
            project_code = row.get("project_code")
            entry_date = row.get("date")
            hours = row.get("hours")

            if not username or not project_code or not entry_date or not hours:
                continue

            user = self.refs["users"].get(username)
            if not user:
                continue

            from apps.projects.models import Project

            try:
                project = Project.objects.get(tenant=self.tenant, code=project_code)
            except Project.DoesNotExist:
                continue

            phase_ref = row.get("phase_ref")
            phase = self.refs["phases"].get(f"{project_code}:{phase_ref}")

            # Handle date conversion
            if isinstance(entry_date, str):
                entry_date = date.fromisoformat(entry_date)

            from apps.time_entries.models import TimeEntry

            TimeEntry.objects.get_or_create(
                tenant=self.tenant,
                employee=user,
                project=project,
                phase=phase,
                date=entry_date,
                defaults={
                    "hours": float(hours),
                    "notes": row.get("notes", ""),
                    "status": row.get("status", "DRAFT"),
                },
            )
            count += 1

        self.stdout.write(self.style.SUCCESS(f"  ✓ {count} entrées de temps importées"))


class DryRunRollback(Exception):
    pass
